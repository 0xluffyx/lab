#!/usr/bin/env python3
"""
企查查批量查询社保参保人数
用法: python3 qcc_batch_query.py <excel文件>

前置:
  pip install openpyxl requests
  export QCC_API_KEY='Bearer 你的key'
  在国内网络环境下运行
"""
import os, sys, json, time
import requests
import openpyxl

API_KEY = os.environ.get("QCC_API_KEY", "")
if not API_KEY:
    print("请先设置: export QCC_API_KEY='Bearer 你的key'")
    sys.exit(1)

MCP_URL = "https://agent.qcc.com/mcp/company/stream"

def query_company(name):
    """通过 MCP API 查询企业工商详情，提取参保人数"""
    headers = {
        "Authorization": API_KEY,
        "Content-Type": "application/json",
        "Accept": "text/event-stream",
    }
    payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {
            "name": "get_company_registration_info",
            "arguments": {"searchKey": name}
        }
    }
    try:
        resp = requests.post(MCP_URL, headers=headers, json=payload, timeout=15)
        if resp.status_code != 200:
            print(f"  HTTP {resp.status_code}: {resp.text[:200]}")
            return ""
        
        # 解析 SSE 响应
        for line in resp.text.split('\n'):
            if line.startswith('data:'):
                try:
                    data = json.loads(line[5:])
                    if 'result' in data:
                        content = data['result'].get('content', [])
                        for item in content:
                            if item.get('type') == 'text':
                                # 尝试解析文本中的 JSON
                                text = item['text']
                                # 递归查找 insuredCount 字段
                                def find_insured(obj):
                                    if isinstance(obj, dict):
                                        for k, v in obj.items():
                                            if 'insured' in str(k).lower() or '社保' in str(k) or '参保' in str(k):
                                                return v
                                            r = find_insured(v)
                                            if r is not None:
                                                return r
                                    elif isinstance(obj, list):
                                        for item in obj:
                                            r = find_insured(item)
                                            if r is not None:
                                                return r
                                    return None
                                count = find_insured(json.loads(text) if text.strip().startswith('{') or text.strip().startswith('[') else {})
                                if count is not None:
                                    return str(count)
                                # 如果解析失败，尝试从文本中直接匹配
                                import re
                                m = re.search(r'参保人数[：:]\s*(\d+)', text)
                                if m:
                                    return m.group(1)
                except:
                    continue
    except Exception as e:
        print(f"  查询失败: {e}")
    return ""

def main():
    f = sys.argv[1]
    if not os.path.exists(f):
        print(f"文件不存在: {f}")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column+1)}
    nc, ic = hdr["企业名称"], hdr["社保人数"]
    
    total = ws.max_row - 1
    print(f"共 {total} 家企业\n")
    
    for i in range(2, ws.max_row + 1):
        name = ws.cell(i, nc).value
        if not name:
            continue
        # 已有数据跳过
        if ws.cell(i, ic).value and str(ws.cell(i, ic).value).strip():
            print(f"[{i-1}/{total}] {name} → 已有: {ws.cell(i, ic).value}")
            continue
        
        print(f"[{i-1}/{total}] 查询: {name}...", end=" ")
        cnt = query_company(str(name).strip())
        ws.cell(i, ic, value=cnt)
        wb.save(f)
        print(f"✅ {cnt}" if cnt else "⚠️ 未获取到")
        time.sleep(1.5)  # 控制频率
    
    print(f"\n✅ 完成! 结果已保存到: {f}")

if __name__ == "__main__":
    main()
